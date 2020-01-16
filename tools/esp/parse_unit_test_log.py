import argparse
import os
import re
import openpyxl
# import pandas as pd



UNIT_LOG_FILE = r"C:\proj\test\utest_script\dual_ttyUSB3.md"
UNIT_RX_LOG_FILE = r"C:\proj\test\utest_script\dual_ttyUSB1.md"

PATTERN_ONE_CASE = re.compile(r"Running ([\s\S]+?)\.\.\.([\s\S]+?)PASS")
PATTERN_RX_CASE = re.compile(r"Running ([\s\S]+?)(?:,|\.\.\.))([\s\S]+?)LAST_END")

# PATTERN_FAIL_PERCENT = re.compile(r"fail:(\d+\(\d+.\d+%\))")
# PATTERN_COUNTS = re.compile(r"((?:enable:\d+)|(?:complete:\d+/?\d{0,10})|(?:success:\d+)|(?:rtt\(max:\d+, min:\d+\)))")
# PATTERN_COUNTS = re.compile(r"((?:enable:\d+)|(?:complete:\d+/?\d{0,10})|(?:success:\d+))")


PATTERNS = [
# RX
    ["RX ID",                re.compile(r"PSDU\(id:(\d+), ssn:\d+\)")],
    ["(ID Lost)",            re.compile(r"id_lost:(\d+)")],       #add "ID Lost" column
    ["TX ID",                re.compile(r"psdu\(id:(\d+), ssn:\d+\)")],
    ["RX SSN",               re.compile(r"PSDU\(id:\d+, ssn:(\d+)\)")],
    ["(SSN Lost)",           re.compile(r"ssn_lost:(\d+)")],      #add "SSN Lost" column
    ["TX SSN",               re.compile(r"psdu\(id:\d+, ssn:(\d+)\)")],
    ["RX Success",           re.compile(r"PASS, rx count:(\d+)")],
    ["(DIFF/SUC)",           re.compile(r"rx/tx:(\d+)")],         #add "DIFF/SUC" column, compare "RX Success" and "TX Success".
    ["TX Min Bitmap",        re.compile(r"min_bitmap:(\d+)")],
    ["TX Max Bitmap",        re.compile(r"max_bitmap:(\d+)")],
    ["TX Success",           re.compile(r"success:(\d+)")],
    ["RX MPDUs",             re.compile(r"WDEVRX_MPDU_CNT           :(\d+)")],
    ["RX BFULL",             re.compile(r"WDEVRX_BUF_FULLCNT        :(\d+)")],
    ["(RX/DIFF/MPDUs)",      re.compile(r"(rx/diff/mpdus):(\d+)")],  #add "DIFF/MPDUs" column, RX side: "RX MPDUs" = "RX Success" + "RX BFULL"
    ["(Lost/MPDUs)",         re.compile(r"(lost/mpdus):(\d+)")],     #add "Lost/MPDUs" column, compare "RX MPDUs" and "TX MPDUs".
    ["TX MPDUs",             re.compile(r"mpdu:(\d+)")],
    ["RX DATA",              re.compile(r"WDEVRX_DATASUC_CNT        :(\d+)")],
    ["RX END",               re.compile(r"WDEVRX_END_CNT            :(\d+)")],
    ["(RX/DIFF/ISR)",        re.compile(r"(rx/diff/isr) :(\d+)")],   #add "DIFF/MPDUs" column, RX side: "RX END" = "RX DATA" + "RX BFULL"
    ["(DIFF/END)",           re.compile(r"rx_end/tx_end:(\d+)")],    #add "DIFF/END" column, compare "RX END"(rx isr) and "Enable"(tx isr).
    ["Enable",               re.compile(r"enable:(\d+)")],
    ["Complete",             re.compile(r"complete:(\d+)")],
    ["RX FCS ERR",           re.compile(r"WDEVRX_FCS_ERRCNT         :(\d+)")],
    ["TX FCS ERR",           re.compile(r"WDEVRX_TX__FCS_ERRCNT         :(\d+)")],
    ["RX ACK INT",           re.compile(r"WDEVRXACK_INT_CNT         :(\d+)")],
    ["TX ACK INT",           re.compile(r"WDEVRX_TX_ACK_INT_CNT         :(\d+)")],

# TX
    ["Throughput",       re.compile(r"(\d+.\d+) Mbps")],
    ["Lost Percent(%)",  re.compile(r"fail:\d+\((\d+.\d+)%\)")],
    ["Lost Packets",     re.compile(r"fail:(\d+)\(\d+.\d+%\)")],
    ["(DIFF/BFLL)",      re.compile(r"diff/bfull:(\d+)")],        #add "DIFF/BFULL" column, confirm if the "Lost Ba/Ack" is caused by "RX BFULL".
    ["Lost Ba/Ack",      re.compile(r"lost:(\d+)")],
    ["Fail Reason",      re.compile(r"\(test\)\[\d+\]\[\d+\]\[\d+\]([\s\S]+?[ ]{0,3}\d+/[ ]{0,3}\d+\([ ]{0,3}\d+.\d+%\))")],
    ["Max RTT",          re.compile(r"max_rtt:(\d+)")],
    ["Min RTT",          re.compile(r"min_rtt:(\d+)")],
    ["(DIFF/RTT)",       re.compile(r"diff/rtt:(\d+)")],          #add "DIFF/RTT" column, compare "Max RTT" and "Min RTT".

# TX
    ["TX MPDU CNT",             re.compile(r"WDEVRX_TX__MPDU_CNT           :(\d+)")],
    ["TX BFULL",                re.compile(r"WDEVRX_TX__BUF_FULLCNT        :(\d+)")],
    ["TX RX DATA",              re.compile(r"WDEVRX_TX__DATASUC_CNT        :(\d+)")],
    ["TX RX END",               re.compile(r"WDEVRX_TX__END_CNT            :(\d+)")],

# RX
    ["RX CCK ERR",              re.compile(r"WDEVRX_CCK_ERRCNT         :(\d+)")],
    ["RX OFDM ERR",             re.compile(r"WDEVRX_OFDM_ERRCNT        :(\d+)")],
    ["RX AGC ERR",              re.compile(r"WDEVRX_AGC_ERRCNT         :(\d+)")],
    ["RX Shorten",              re.compile(r"WDEVRX_SF_CNT             :(\d+)")],
    ["RX Abort",                re.compile(r"WDEVRX_ABORT_CNT          :(\d+)")],
    ["RX FIFO OVFL",            re.compile(r"WDEVRX_FIFO_OVFCNT        :(\d+)")],
    ["RX APENTRYBUF FULL",      re.compile(r"WDEVRX_APENTRYBUF_FULLCNT :(\d+)")],
    ["RX OTHER UCAST",          re.compile(r"WDEVRX_OTHER_UNICASTCNT   :(\d+)")],
    ["RX TKIP ERR",             re.compile(r"WDEVRX_TKIP_ERRCNT        :(\d+)")],
    ["RX SAME BITMP",           re.compile(r"WDEVRX_SAMEBM_ERRCNT      :(\d+)")],
    ["RX RTS INT",              re.compile(r"WDEVRXRTS_INT_CNT         :(\d+)")],
    ["RX RIFS INT",             re.compile(r"WDEVRXRIFS_INT_CNT        :(\d+)")],
    ["RX BT BLOCK ERR",         re.compile(r"WDEVRX_BTBLOCK_ERR_CNT    :(\d+)")],
    ["RX FREQ HOP ERR",         re.compile(r"WDEVRX_FREQHOP_ERR_CNT    :(\d+)")],
    ["RX LAST UNMATCH",         re.compile(r"WDEVRX_LASTUNMATCH_ERR_CNT:(\d+)")],
    ["RX BLOCK ERR",            re.compile(r"WDEVRX_BLOCK_ERR_CNT      :(\d+)")],

# TX
    ["TX CCK ERR",              re.compile(r"WDEVRX_TX__CCK_ERRCNT         :(\d+)")],
    ["TX OFDM ERR",             re.compile(r"WDEVRX_TX__OFDM_ERRCNT        :(\d+)")],
    ["TX AGC ERR",              re.compile(r"WDEVRX_TX__AGC_ERRCNT         :(\d+)")],
    ["TX Shorten",              re.compile(r"WDEVRX_TX__SF_CNT             :(\d+)")],
    ["TX Abort",                re.compile(r"WDEVRX_TX__ABORT_CNT          :(\d+)")],
    ["TX FIFO OVFL",            re.compile(r"WDEVRX_TX__FIFO_OVFCNT        :(\d+)")],
    ["TX APENTRYBUF FULL",      re.compile(r"WDEVRX_TX__APENTRYBUF_FULLCNT :(\d+)")],
    ["TX OTHER UCAST",          re.compile(r"WDEVRX_TX__OTHER_UNICASTCNT   :(\d+)")],
    ["TX TKIP ERR",             re.compile(r"WDEVRX_TX__TKIP_ERRCNT        :(\d+)")],
    ["TX SAME BITMP",           re.compile(r"WDEVRX_TX__SAMEBM_ERRCNT      :(\d+)")],
    ["TX RTS INT",              re.compile(r"WDEVRX_TX_RTS_INT_CNT         :(\d+)")],
    ["TX RIFS INT",             re.compile(r"WDEVRX_TX_RIFS_INT_CNT        :(\d+)")],
    ["TX BT BLOCK ERR",         re.compile(r"WDEVRX_TX__BTBLOCK_ERR_CNT    :(\d+)")],
    ["TX FREQ HOP ERR",         re.compile(r"WDEVRX_TX__FREQHOP_ERR_CNT    :(\d+)")],
    ["TX LAST UNMATCH",         re.compile(r"WDEVRX_TX__LASTUNMATCH_ERR_CNT:(\d+)")],
    ["TX BLOCK ERR",            re.compile(r"WDEVRX_TX__BLOCK_ERR_CNT      :(\d+)")],
]


def write_excel_xls(xls_file, result):
    if os.path.exists(xls_file):
        os.remove(xls_file)
    workbook = openpyxl.Workbook()
    # sheet = workbook.create_sheet("Sheet1")
    sheet = workbook.active
    # headers = result[0].keys()
    headers = ["Case name"]
    headers.extend([key for key,_ in PATTERNS])
    for j, title in enumerate(headers):
        # sheet.write(0, j, title)
        sheet.cell(row=1, column=j+1).value = title
    for i, line in enumerate(result):
        for j, key in enumerate(headers):
            try:
                # sheet.write(i+1, j, float(line[key]))
                sheet.cell(row=i+2, column=j+1).value = float(line[key])
            except ValueError:
                # sheet.write(i+1, j, line[key])
                sheet.cell(row=i+2, column=j+1).value = line[key]
    workbook.save(xls_file)

def parse_unit_log(log_file, rx_log_file=""):
    with open(log_file, "rb") as f:
        all_log = f.read()
        all_log = all_log.replace("Running tests matching", "----ignore---")
        all_log = re.sub(r"ampdu_count:\d+, psdu\(id:\d+, ssn:\d+\), seqno:\d+,", "", all_log)
        all_log = re.sub(r"esp_test_tx_process_complete,\d+\] \(test\)aci:\d+, lost Ba/Ack, ampdu_count:\d+, seqno:\d+, psdu\(id:\d+, ssn:\d+\),", "", all_log)
        all_log = all_log.replace("WDEVRX", "WDEVRX_TX_")
    cases = PATTERN_ONE_CASE.finditer(all_log)
    # cases = PATTERN_ONE_CASE.findall(all_log)

    if rx_log_file:
        with open(rx_log_file, "rb") as f:
            all_rx_log = f.read()
            all_rx_log = all_rx_log.replace("Running tests matching", "----ignore---")
            all_rx_log = all_rx_log.replace("Running RX MAC...", "Running TX MAC...")
            all_rx_log = all_rx_log.replace("Running Set MAC RX", "Running Set MAC TX")
            all_rx_log = all_rx_log.replace("(test)rx, Running", "LAST_END.\n (test)rx, Running")
        rx_cases = PATTERN_RX_CASE.findall(all_rx_log)
    else:
        rx_cases = []

    # print("tx case", ",".join(case[0] for case in cases))
    # print("rx_case", ",".join(case[0] for case in rx_cases))
    # print(len(cases))
    # print(len(rx_cases))

    results = []
    for case in cases:
        name = case.group(1)
        tx_data = case.group(2)

        # Assume rx_cases is less than tx_cases
        # always uses the first match in rx_cases here
        if rx_cases and rx_cases[0][0] == name:
            rx_data = rx_cases[0][1]
            rx_cases = rx_cases[1:]
        else:
            if rx_log_file:
                print("There's no such case in rx log, case_name:{}".format(name))
            rx_data = ""

        data = tx_data + rx_data
        parsed_case = {"Case name": name}
        for key,pattern in PATTERNS:
            matchs = pattern.findall(data)
            parsed_case.update({key: "\n".join(matchs)})
        results.append(parsed_case)

    if rx_cases:
        print("extra rx cases:")
        print(",".join([c[0] for c in rx_cases]))

    xls_file = os.path.join(os.path.dirname(log_file),"RESULT.xls")
    write_excel_xls(xls_file, results)

    

def main():
    log_file = ''
    try: 
        parser = argparse.ArgumentParser()
        parser.add_argument("log_file")
        parser.add_argument("-r", "--rx_log_file", help="utest rx log file", default="")
        args = parser.parse_args()
        log_file = args.log_file
        rx_log_file = args.rx_log_file
    except SystemExit:
        log_file = UNIT_LOG_FILE
        rx_log_file = UNIT_RX_LOG_FILE
        print("try to parse {}".format(UNIT_LOG_FILE))

    parse_unit_log(log_file, rx_log_file)

if __name__ == "__main__":
    main()
